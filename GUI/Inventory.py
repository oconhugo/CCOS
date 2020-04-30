import os

import xlsxwriter
import openpyxl as op
from datetime import date

today = date.today()
today_formated = today.strftime("%b-%d-%Y")
current_dir = os.getcwd()
inventory_dir = current_dir + "/Inventory"
excel_dir = inventory_dir + "/" + today_formated + ".xlsx"

if not os.path.exists(inventory_dir):
    os.makedirs(inventory_dir)

workbook = xlsxwriter.Workbook(excel_dir)
worksheet = workbook.add_worksheet("Inventario")
row , orderNum = 0, 0
counter = 0
date = ""

class Inv(object):
    def __init__(self,master,**kwargs):
        # Create a workbook and add a worksheet.
        global date
        date = today_formated

    def add(self, order):
        global row, orderNum
        orderNum += 1
        cell_header_format = workbook.add_format()
        cell_header_format.set_bold()
        cell_header_format.set_bg_color('black')
        cell_header_format.set_font_color('white')
        cell_header_format.set_left()
        cell_header_format.set_right()
        cell_header_format.set_left_color('white')
        cell_header_format.set_right_color('white')

        cell_header_format_1 = workbook.add_format()
        cell_header_format_1.set_bold()
        cell_header_format_1.set_bg_color('black')
        cell_header_format_1.set_font_color('white')
        cell_header_format_1.set_border()
        cell_header_format_1.set_border_color('white')

        worksheet.set_column(0, 0, 15)
        worksheet.set_column(1, 1, 30)
        worksheet.set_column(2, 2, 5)
        worksheet.set_column(5, 5, 25)
        worksheet.set_column(6, 6, 10)

        if not os.path.exists(excel_dir):
            orderNum=1
            worksheet.write(row, 0, "Order #", cell_header_format)
            worksheet.write(row, 1, "Items", cell_header_format)
            worksheet.write(row, 2, "Cost", cell_header_format)
            worksheet.write(0, 5, "Total Number of Orders: ", cell_header_format_1)
            worksheet.write(1, 5, "Total Earnings: ", cell_header_format_1)
            worksheet.write(0, 6, "=COUNTA(A:A)-1")
            worksheet.write(1, 6, "=SUM(C:C)")

            row+=1
            colText =0
            worksheet.write(row, colText, "Order #"+ str(orderNum))
            row+=1
            colText, colValue = 1, 2
            for i in order:
                for key in i:
                    worksheet.write(row, colText, key)
                    worksheet.write(row, colValue, i[key])
                row += 1
                orderNum=+1
            Inv.closeXLSX(self)
        else:
            wb = op.load_workbook(excel_dir)
            ws = wb.get_sheet_by_name('Inventario')
            row += 1
            empty = ""
            ws.append(["Order #" + str(orderNum)])
            for i in order:
                for key in i:
                    ws.append([empty, key, i[key]])
            wb.save(excel_dir)
            wb.close()

    def closeXLSX(self):
        global date
        workbook.close()
        #ss = openpyxl.load_workbook(excel_dir)
        #ss_sheet = ss['Sheet1']
        #ss_sheet.title = date
        #ss.save(excel_dir)